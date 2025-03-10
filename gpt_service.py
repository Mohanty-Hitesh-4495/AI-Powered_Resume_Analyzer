from openai import OpenAI
import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv
from ocr_service import analyze_read_return
import os
import json

# loading environment variables
load_dotenv()
openai_api_key = os.getenv("OPENAI_API_KEY")

openai = OpenAI(
    api_key = openai_api_key
)

def extract_details_with_gpt(resume_text):
    # defining the GPT prompt
    prompt = [
        {
            "role": "system",
            "content": "You are a helpful assistant that extracts structured information from resumes."
        },
        {
            "role": "user",
            "content": f"""
            Extract the following details from the given resume text:

            Mandatory Fields:
            - Name
            - Contact details (email, phone number)
            - University
            - Year of Study
            - Course
            - Discipline
            - CGPA/Percentage
            - Key Skills
            - Gen AI Experience Score (Keep it empty)
            - AI/ML Experience Score (Keep it empty)
            - Overall Score (Keep it empty)
            - Supporting Information (2 most highlighting informations e.g., certifications, internships, projects)

            Resume Text:
            {resume_text}

            Output JSON Format:
            {{
            "Name": "",
            "Email": "",
            "Phone Number": "",
            "University": "",
            "Year Of Study": "",
            "Course": "",
            "Discipline": "",
            "CGPA": "",
            "Key Skills": "",
            "Gen-AI Experience Score": 0,
            "AI/ML Experience Score": 0,
            "Overall Score": 0,
            "Supporting Information": ""
            }}
            """
        }
    ]

    # calling GPT model
    response = openai.chat.completions.create(
        model="gpt-4o-mini",  
        messages=prompt,
        temperature=0.5
    )
    return response.choices[0].message.content.strip()

# OpenAI API scoring function
def evaluate_resume(resume_text):
    messages = [
        {
            "role": "system",
            "content": "You are an expert evaluator assessing resumes for AI/ML and Gen AI expertise."
        },
        {
            "role": "user",
            "content": f"""
    Based on the scoring system below, evaluate the following resume and return a concise JSON response:

    Resume: {resume_text}

    Scoring System:
    - Gen AI Experience Score:
    0 (Unexposed): No exposure or experience with Gen AI concepts or tools.
    1 (Exposed): Basic understanding or exposure to foundational Gen AI concepts (e.g., GPT basics, etc).
    2 (Hands-on): Practical experience working on tasks involving prompt engineering, fine-tuning, or embeddings.
    3 (Advanced): Advanced work such as Agentic RAG, evaluation frameworks (Evals), or building agent-based systems.

    - AI/ML Experience Score:
    0 (Unexposed): No exposure or experience with AI/ML concepts or tools.
    1 (Exposed): Basic understanding of AI/ML concepts (e.g., regression, classification, etc).
    2 (Hands-on): Implemented projects using ML/DL frameworks (e.g., TensorFlow, PyTorch, etc).
    3 (Advanced): Published research, implemented innovative models, or production-level deployment of ML/DL solutions.

    Return the response as a JSON object with the following structure:
    {{
    "gen_ai_experience_score": {{
        "score": <integer between 0-3>
    }},
    "ai_ml_experience_score": {{
        "score": <integer between 0-3>
    }},
    "overall_score": {{
        "score": <float weighted average of AI/ML and Gen-AI Scores between 0-3 with 2 decimal>
    }}
    }}
    """
        }
    ]

    # clling OpenAI API
    response = openai.chat.completions.create(
        model="gpt-4o-mini",
        messages=messages,
    )

    return response.choices[0].message.content.strip()

# function to append data to an existing Excel file
def append_to_excel(file_path, data, header):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # appending each row
        for row in data:
            row = [', '.join(element) if isinstance(element, list) else element for element in row]
            sheet.append(row)
        
        # adjusting column widths
        for col_idx, col_name in enumerate(header, start=1):
            max_len = max(
                len(str(col_name)),  # length of header
                max(
                    len(str(sheet.cell(row=row_idx, column=col_idx).value or ""))  # length of cell values
                    for row_idx in range(1, sheet.max_row + 1)
                ) 
            ) + 2  # padding
            sheet.column_dimensions[sheet.cell(row=1, column=col_idx).column_letter].width = max_len

        workbook.save(file_path)
        print(f"Data appended to {file_path}.")

    except FileNotFoundError:
        # if the file does not exist, create it
        df = pd.DataFrame(data, columns=header)
        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')

            # adjusting column widths for the new file
            worksheet = writer.sheets['Sheet1']
            for idx, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.set_column(idx, idx, max_len)
        print(f"New file created at {file_path} with adjusted column spacing.")

if __name__ == "__main__":
    path_to_doc = "sample_resumes/Resume-Syed Sadiqu Hussain.pdf"
    with open(path_to_doc, "rb") as f:
        extracted_text = analyze_read_return(f)

    # extracting details using GPT
    extracted_details = extract_details_with_gpt(extracted_text)

    """ generating excel for updating data extracted from each resume """
    try:
        # cleaning and parsing the JSON output from GPT
        cleaned_details = extracted_details.strip("```json").strip("```").strip()
        details_dict = json.loads(cleaned_details)

        # evaluating the resume for scores
        scores_response = evaluate_resume(extracted_text)
        clean_scores = scores_response.strip("```json").strip("```").strip()

        # parsing the score response and update details
        scores = json.loads(clean_scores)  
        print(scores)
        details_dict["Gen-AI Experience Score"] = scores.get("gen_ai_experience_score", {}).get("score", 0)
        details_dict["AI/ML Experience Score"] = scores.get("ai_ml_experience_score", {}).get("score", 0)
        details_dict["Overall Score"] = scores.get("overall_score", {}).get("score",0)

        df = pd.DataFrame([details_dict])

        # saving updated details to an Excel file
        output_path = 'parsed_resumes.xlsx'
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Resume Details', index=False)
            worksheet = writer.sheets['Resume Details']

            # Adjust column widths based on content
            for idx, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.set_column(idx, idx, max_len)
        print(f"Details successfully saved to {output_path}")
        
    except json.JSONDecodeError as e:
        print(f"Failed to parse JSON: {e}")